import os
import openpyxl
import sqlite3
from datetime import datetime
from PyQt6.QtCore import QObject, pyqtSignal
from core.db_manager_utils import BASE_DIR_PATH, get_yearly_db_path, GLOBAL_DB_PATH

class DataMigrator(QObject):
    progress_changed = pyqtSignal(int, int, str)  # current, total, label
    log_message = pyqtSignal(str, bool, str)      # message, is_error, target
    finished = pyqtSignal(int)
    error_occurred = pyqtSignal(str)
    request_confirmation = pyqtSignal(int, str) # year, message. To be handled by UI.

    def __init__(self):
        super().__init__()
        self._cancel_requested = False
        self._confirmation_result = None

    def cancel(self):
        self._cancel_requested = True

    def set_confirmation_result(self, result: bool):
        self._confirmation_result = result

    def migrate_resources(self):
        if not os.path.exists(BASE_DIR_PATH):
            self.error_occurred.emit(f"Ruta base {BASE_DIR_PATH} no encontrada.")
            return

        years = list(range(2004, datetime.now().year + 1))
        total_migrated = 0
        type_res_map = {}
        seasons_map = {}

        # Use sqlite3 standard lib in background thread
        try:
            conn_global = sqlite3.connect(GLOBAL_DB_PATH)
            cursor_global = conn_global.cursor()

            cursor_global.execute("SELECT idx, type_resource FROM T_Type_Resources")
            for row in cursor_global.fetchall():
                type_res_map[row[1]] = row[0]

            cursor_global.execute("SELECT precure_season_name FROM T_Seasons")
            for row in cursor_global.fetchall():
                seasons_map[row[0]] = row[0]

            conn_global.close()
        except Exception as e:
            self.error_occurred.emit(f"Error accessing global DB: {e}")
            return

        for i, year in enumerate(years):
            if self._cancel_requested:
                self.log_message.emit("Migración cancelada por el usuario.", True, "resources")
                break

            self.progress_changed.emit(i, len(years), f"Procesando año {year}...")
            self.log_message.emit(f"Iniciando migración de recursos para el año {year}...", False, "resources")

            px = year - 2003
            px_str = f"{px:02d}"
            excel_path = os.path.join(BASE_DIR_PATH, str(year), f"{px_str}. identity_propeties", f"{px_str}. le_etude.overwrite.xlsx")

            if not os.path.exists(excel_path):
                continue

            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                if "material_list" not in wb.sheetnames:
                    continue

                sheet = wb["material_list"]
                db_year_path = get_yearly_db_path(year)
                conn_year = sqlite3.connect(db_year_path)
                cursor_year = conn_year.cursor()

                existing_titles = set()
                cursor_year.execute("SELECT title_material FROM T_Resources")
                for row in cursor_year.fetchall():
                    existing_titles.add(row[0])
                for row_idx in range(4, sheet.max_row + 1):
                    if self._cancel_requested: break
                    
                    title_material = sheet.cell(row=row_idx, column=9).value
                    if not title_material: continue

                    base_title = str(title_material)
                    final_title = base_title
                    counter = 2
                    while final_title in existing_titles:
                        final_title = f"{base_title} ({counter})"
                        counter += 1

                    existing_titles.add(final_title)
                    type_mat_id = type_res_map.get(sheet.cell(row=row_idx, column=5).value)
                    season_name_fk = seasons_map.get(sheet.cell(row=row_idx, column=6).value)

                    try:
                        cursor_year.execute("""
                            INSERT INTO T_Resources (
                                title_material, type_material, precure_season_name, ep_num, ep_sp_num,
                                released_utc_09, released_soundtrack_utc_09, released_spinoff_utc_09,
                                duration_file, datetime_download, relative_path_of_file,
                                relative_path_of_soundtracks, relative_path_of_lyrics
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            final_title, type_mat_id, season_name_fk,
                            sheet.cell(row=row_idx, column=7).value,
                            sheet.cell(row=row_idx, column=8).value,
                            str(sheet.cell(row=row_idx, column=10).value or ""),
                            str(sheet.cell(row=row_idx, column=11).value or ""),
                            str(sheet.cell(row=row_idx, column=12).value or ""),
                            str(sheet.cell(row=row_idx, column=13).value or ""),
                            str(sheet.cell(row=row_idx, column=14).value or ""),
                            str(sheet.cell(row=row_idx, column=15).value or ""),
                            None, None
                        ))
                        total_migrated += 1
                        self.log_message.emit(f"Migrado recurso: {final_title}", False, "resources")
                    except Exception as e:
                        self.log_message.emit(f"Error al migrar recurso {final_title}: {e}", True, "resources")

                conn_year.commit()
                conn_year.close()
            except Exception as e:
                self.log_message.emit(f"Error procesando {excel_path}: {e}", True, "resources")

        self.progress_changed.emit(len(years), len(years), "Finalizado.")
        self.finished.emit(total_migrated)

    def migrate_registry(self):
        from core.db_manager_utils import calculate_lapsed, get_opener_model_info_sqlite
        if not os.path.exists(BASE_DIR_PATH):
            self.error_occurred.emit(f"Ruta base {BASE_DIR_PATH} no encontrada.")
            return

        years = list(range(2004, datetime.now().year + 1))
        total_migrated = 0
        for i, year in enumerate(years):
            if self._cancel_requested: break

            self.progress_changed.emit(i, len(years), f"Procesando año {year}...")
            self.log_message.emit(f"Iniciando migración de registros para el año {year}...", False, "registry")

            px = year - 2003
            px_str = f"{px:02d}"
            excel_path = os.path.join(BASE_DIR_PATH, str(year), f"{px_str}. identity_propeties", f"{px_str}. le_etude.overwrite.xlsx")
            if not os.path.exists(excel_path): continue

            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                if "overwrite_registry" not in wb.sheetnames: continue
                sheet = wb["overwrite_registry"]

                db_year_path = get_yearly_db_path(year)
                conn_year = sqlite3.connect(db_year_path)
                cursor_year = conn_year.cursor()

                # Check if table has data
                cursor_year.execute("SELECT COUNT(*) FROM T_Registry")
                has_data = cursor_year.fetchone()[0] > 0

                if has_data:
                    self._confirmation_result = None
                    self.request_confirmation.emit(year, f"La tabla T_Registry del año {year} contiene datos. ¿Desea borrarlos y migrar?")
                    
                    import time
                    from PyQt6.QtCore import QCoreApplication
                    while self._confirmation_result is None:
                        # processEvents is okay here as it's the main thread waiting on a signal
                        # OR if it's the worker thread, it should wait for a signal result
                        time.sleep(0.01)
                        if self._cancel_requested: break
                    
                    if not self._confirmation_result:
                        conn_year.close()
                        continue

                cursor_year.execute("DELETE FROM T_Registry")

                for row_idx in range(4, sheet.max_row + 1):
                    if self._cancel_requested: break
                    title_material = sheet.cell(row=row_idx, column=11).value
                    if not title_material: continue

                    dt_range = sheet.cell(row=row_idx, column=12).value
                    type_repeat = sheet.cell(row=row_idx, column=13).value
                    type_listen = sheet.cell(row=row_idx, column=14).value
                    model_writer = sheet.cell(row=row_idx, column=15).value

                    lapsed = calculate_lapsed(dt_range)
                    op_model, op_name = get_opener_model_info_sqlite(dt_range, model_writer)

                    try:
                        cursor_year.execute("""
                            INSERT INTO T_Registry (
                                title_material, datetime_range_utc_06, type_repeat,
                                type_listen, model_writer, lapsed_calculated,
                                opener_model, name_of_opener_model
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            str(title_material), str(dt_range or ""),
                            str(type_repeat or ""), str(type_listen or ""),
                            str(model_writer or ""), lapsed, op_model, op_name
                        ))
                        total_migrated += 1
                        self.log_message.emit(f"Migrado registro: {title_material} ({dt_range})", False, "registry")
                    except Exception as e:
                        self.log_message.emit(f"Error al migrar registro {title_material}: {e}", True, "registry")

                conn_year.commit()
                conn_year.close()
            except Exception as e:
                self.log_message.emit(f"Error registry {year}: {e}", True, "registry")

        self.progress_changed.emit(len(years), len(years), "Finalizado.")
        self.finished.emit(total_migrated)
