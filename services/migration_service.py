import os
import openpyxl
from datetime import datetime

class MigrationService:
    def __init__(self, catalog_repo, resources_repo, registry_repo):
        self.catalog_repo = catalog_repo
        self.resources_repo = resources_repo
        self.registry_repo = registry_repo
        self._cancel_requested = False

    def cancel(self): self._cancel_requested = True

    def migrate_resources(self, base_dir, years, progress_callback=None, log_callback=None):
        total_migrated = 0
        type_res_map = self.catalog_repo.get_all_resource_types()
        seasons_map = self.catalog_repo.get_all_season_names()

        for i, year in enumerate(years):
            if self._cancel_requested: break
            if progress_callback: progress_callback(i, len(years), f"Procesando año {year}...")
            if log_callback: log_callback(f"Iniciando migración de recursos para el año {year}...", False, "resources")

            px = year - 2003
            px_str = f"{px:02d}"
            excel_path = os.path.join(base_dir, str(year), f"{px_str}. identity_propeties", f"{px_str}. le_etude.overwrite.xlsx")
            if not os.path.exists(excel_path): continue

            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                if "material_list" not in wb.sheetnames: continue
                sheet = wb["material_list"]

                existing_titles = self.resources_repo.get_all_titles()
                for row_idx in range(4, sheet.max_row + 1):
                    if self._cancel_requested: break
                    title_material = sheet.cell(row=row_idx, column=9).value
                    if not title_material: continue

                    base_title = str(title_material)
                    final_title, counter = base_title, 2
                    while final_title in existing_titles:
                        final_title = f"{base_title} ({counter})"
                        counter += 1
                    existing_titles.add(final_title)

                    type_mat_id = type_res_map.get(sheet.cell(row=row_idx, column=5).value)
                    season_name_fk = seasons_map.get(sheet.cell(row=row_idx, column=6).value)

                    data = (
                        final_title, type_mat_id, season_name_fk,
                        sheet.cell(row=row_idx, column=7).value,
                        sheet.cell(row=row_idx, column=8).value,
                        str(sheet.cell(row=row_idx, column=10).value or ""),
                        str(sheet.cell(row=row_idx, column=11).value or ""),
                        str(sheet.cell(row=row_idx, column=12).value or ""),
                        str(sheet.cell(row=row_idx, column=13).value or ""),
                        str(sheet.cell(row=row_idx, column=14).value or ""),
                        str(sheet.cell(row=row_idx, column=15).value or ""),
                        None, None
                    )
                    if self.resources_repo.insert_resource(data):
                        total_migrated += 1
                        if log_callback: log_callback(f"Migrado recurso: {final_title}", False, "resources")
            except Exception as e:
                if log_callback: log_callback(f"Error procesando {excel_path}: {e}", True, "resources")
        return total_migrated

    def migrate_registry(self, base_dir, years, confirm_callback=None, progress_callback=None, log_callback=None):
        total_migrated = 0
        from db_manager import calculate_lapsed # Still helper in db_manager for now

        for i, year in enumerate(years):
            if self._cancel_requested: break
            if progress_callback: progress_callback(i, len(years), f"Procesando año {year}...")
            if log_callback: log_callback(f"Iniciando migración de registros para el año {year}...", False, "registry")

            px = year - 2003
            px_str = f"{px:02d}"
            excel_path = os.path.join(base_dir, str(year), f"{px_str}. identity_propeties", f"{px_str}. le_etude.overwrite.xlsx")
            if not os.path.exists(excel_path): continue

            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                if "overwrite_registry" not in wb.sheetnames: continue
                sheet = wb["overwrite_registry"]

                if self.registry_repo.count_registries() > 0:
                    if confirm_callback and not confirm_callback(year): continue
                self.registry_repo.clear_registries()

                for row_idx in range(4, sheet.max_row + 1):
                    if self._cancel_requested: break
                    title_material = sheet.cell(row=row_idx, column=11).value
                    if not title_material: continue

                    dt_range = sheet.cell(row=row_idx, column=12).value
                    type_repeat = sheet.cell(row=row_idx, column=13).value
                    type_listen = sheet.cell(row=row_idx, column=14).value
                    model_writer = sheet.cell(row=row_idx, column=15).value

                    lapsed = calculate_lapsed(dt_range)
                    op_model, op_name = self.catalog_repo.get_opener_model_info(dt_range, model_writer)

                    data = (str(title_material), str(dt_range or ""), str(type_repeat or ""), str(type_listen or ""), str(model_writer or ""), lapsed, op_model, op_name)
                    if self.registry_repo.insert_registry(data):
                        total_migrated += 1
                        if log_callback: log_callback(f"Migrado registro: {title_material} ({dt_range})", False, "registry")
            except Exception as e:
                if log_callback: log_callback(f"Error registry {year}: {e}", True, "registry")
        return total_migrated
